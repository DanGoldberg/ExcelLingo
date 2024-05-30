import argparse
from typing import Dict

from googletrans import LANGUAGES, Translator
from openpyxl import Workbook, load_workbook


class ExcelTranslator:
    def __init__(
        self, source_excel: str, target_excel: str, source_lang: str, target_lang: str
    ) -> None:
        self.source_excel = source_excel
        self.target_excel = target_excel
        self.source_lang = source_lang
        self.target_lang = target_lang
        self.translator = Translator()
        self.cache: Dict[str, str] = {}

    @staticmethod
    def available_languages() -> str:
        return ", ".join([f"{code}: {name}" for code, name in LANGUAGES.items()])

    @staticmethod
    def validate_language(lang: str) -> str:
        if lang not in LANGUAGES:
            available_languages = ExcelTranslator.available_languages()
            raise argparse.ArgumentTypeError(
                f"Invalid language code: {lang}. Please use a valid language code.\n"
                f"Available languages and codes are:\n{available_languages}"
            )
        return lang

    def translate_text(self, text: str) -> str:
        if text in self.cache:
            return self.cache[text]
        try:
            translation = self.translator.translate(
                text, src=self.source_lang, dest=self.target_lang
            ).text
        except TypeError:
            # Keep original text if translation fails
            translation = text

        self.cache[text] = translation
        return translation

    def translate_excel(self):
        wb = load_workbook(filename=self.source_excel)
        ws = wb.active

        translated_wb = Workbook()
        translated_ws = translated_wb.active

        for row in ws.iter_rows(values_only=True):
            translated_row = []
            for cell in row:
                translated_text = self.translate_text(str(cell))
                translated_row.append(translated_text)
            translated_ws.append(translated_row)

        translated_wb.save(self.target_excel)


def parse_arguments() -> argparse.Namespace:
    available_languages = ExcelTranslator.available_languages()
    parser = argparse.ArgumentParser(
        description="Translate an Excel file from source language to target language.",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=f"Available languages and codes are:\n{available_languages}",
    )
    parser.add_argument("source_excel", type=str, help="Path to the source Excel file.")
    parser.add_argument("target_excel", type=str, help="Path to the target Excel file.")
    parser.add_argument(
        "source_lang",
        type=ExcelTranslator.validate_language,
        help='Source language code (e.g., "en" for English).',
    )
    parser.add_argument(
        "target_lang",
        type=ExcelTranslator.validate_language,
        help='Target language code (e.g., "es" for Spanish).',
    )
    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    translator = ExcelTranslator(
        args.source_excel, args.target_excel, args.source_lang, args.target_lang
    )
    translator.translate_excel()


if __name__ == "__main__":
    main()
