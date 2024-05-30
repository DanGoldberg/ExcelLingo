import argparse
import os
import unittest

from openpyxl import Workbook, load_workbook

from excel_translator import ExcelTranslator


class TestExcelTranslator(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        # Initialize test data
        cls.test_data = {"English": ["Hello", "Goodbye"], "Spanish": ["Hola", "Adiós"]}

    def setUp(self):
        # Create a temporary directory for test files
        self.test_dir = "test_files"
        os.makedirs(self.test_dir, exist_ok=True)

    def tearDown(self):
        # Clean up: Remove temporary directory and its contents
        if os.path.exists(self.test_dir):
            for file in os.listdir(self.test_dir):
                file_path = os.path.join(self.test_dir, file)
                os.remove(file_path)
            os.rmdir(self.test_dir)

    def test_translate_text(self):
        translator = ExcelTranslator("test.xlsx", "output.xlsx", "en", "es")

        translated_text = translator.translate_text("Hello")
        self.assertEqual(translated_text, "Hola")

    def test_translate_excel(self):
        test_file = os.path.join(self.test_dir, "test.xlsx")
        output_file = os.path.join(self.test_dir, "output.xlsx")

        workbook = Workbook()
        sheet = workbook.active
        # Add headers
        headers = list(self.test_data.keys())
        sheet.append(headers)
        # Add data
        for row_data in zip(*self.test_data.values()):
            sheet.append(row_data)
        # Save workbook
        workbook.save(test_file)

        # Initialize ExcelTranslator instance
        translator = ExcelTranslator(test_file, output_file, "en", "es")

        # Translate Excel file
        translator.translate_excel()

        # Read translated output
        output_wb = load_workbook(output_file)
        output_sheet = output_wb.active

        # Check if translations match expected output
        output_data = {}
        for col_idx in range(1, output_sheet.max_column + 1):
            column = output_sheet.cell(row=1, column=col_idx).value
            output_data[column] = [
                output_sheet.cell(row=row_idx, column=col_idx).value
                for row_idx in range(2, output_sheet.max_row + 1)
            ]

        self.assertTrue("English" in output_data)
        self.assertTrue("Spanish" in output_data)
        self.assertEqual(output_data["English"], output_data["Spanish"])

    def test_validate_language_valid(self):
        self.assertEqual(ExcelTranslator.validate_language("en"), "en")
        self.assertEqual(ExcelTranslator.validate_language("es"), "es")

    def test_validate_language_invalid(self):
        with self.assertRaises(argparse.ArgumentTypeError):
            ExcelTranslator.validate_language("xyz")


if __name__ == "__main__":
    unittest.main()
